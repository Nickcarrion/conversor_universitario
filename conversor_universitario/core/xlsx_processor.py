import openpyxl
from tabulate import tabulate


def convert_xlsx(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    partes = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        partes.append(f"\n## Hoja: {sheet_name}\n\n")

        filas = []
        for fila in ws.iter_rows(values_only=True):
            if any(celda is not None for celda in fila):
                fila_limpia = [
                    str(c).replace("\n", " ") if c is not None else ""
                    for c in fila
                ]
                filas.append(fila_limpia)

        if filas:
            encabezado = filas[0]
            datos = filas[1:]
            tabla_md = tabulate(datos, headers=encabezado, tablefmt="github")
            partes.append(tabla_md + "\n")
        else:
            partes.append("*Hoja vacia*\n")

    wb.close()
    return "".join(partes)
